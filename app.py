"""
Club Argentino — Dashboard de Análisis
Desarrollado por Ramón Codesido · Sport Data Campus
"""

import base64
import os
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from data_loader import load_all

# ── Colores del club ──────────────────────────────────────────────────────────
AZUL_OSCURO  = "#1B2A4A"
AZUL_CELESTE = "#6AAFE6"
AZUL_MEDIO   = "#2E4B7A"
AZUL_CLARO   = "#3D6B9E"
DORADO       = "#C9A84C"
BLANCO       = "#FFFFFF"
GRIS_MEDIO   = "#C8D6E5"
VERDE        = "#4CAF82"
ROJO         = "#E85D75"

ASSETS = os.path.join(os.path.dirname(__file__), "assets")
TRAMO_ORDER = ["0-15", "16-30", "31-45", "45+", "46-60", "61-75", "76-90", "90+"]


# ── Helper: template de plotly con profundidad ────────────────────────────────
# Color del área de plot ligeramente más claro que el paper para crear profundidad
PLOT_BG  = "#253F6B"   # azul medio más cálido para el área de datos
PAPER_BG = "#1B2A4A"   # azul oscuro del fondo general

def t(fig: go.Figure, height=380, margin=None, **extra) -> go.Figure:
    m = margin or dict(t=36, b=36, l=48, r=36)
    fig.update_layout(
        height=height,
        margin=m,
        paper_bgcolor=PAPER_BG,
        plot_bgcolor=PLOT_BG,
        font=dict(family="Inter, Arial, sans-serif", color=BLANCO, size=12),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=BLANCO, size=11),
                    orientation="h", y=1.08),
        # Borde sutil alrededor del área de plot
        shapes=[dict(
            type="rect", xref="paper", yref="paper",
            x0=0, y0=0, x1=1, y1=1,
            line=dict(color="rgba(106,175,230,0.12)", width=1),
            fillcolor="rgba(0,0,0,0)", layer="above",
        )],
        **extra,
    )
    fig.update_xaxes(
        gridcolor="rgba(200,214,229,0.07)",
        zeroline=False,
        color=BLANCO,
        tickfont=dict(color=GRIS_MEDIO, size=11),
        linecolor="rgba(200,214,229,0.15)",
        linewidth=1,
    )
    fig.update_yaxes(
        gridcolor="rgba(200,214,229,0.07)",
        zeroline=False,
        color=BLANCO,
        tickfont=dict(color=GRIS_MEDIO, size=11),
        linecolor="rgba(200,214,229,0.15)",
        linewidth=1,
    )
    return fig


def depth_bar(color: str, name: str, x, y, text=None, pos="outside",
              secondary_y=False, opacity=1.0) -> go.Bar:
    """Barra con gradiente vertical para dar sensación de profundidad."""
    # Color base y versión más clara para el highlight superior
    return go.Bar(
        name=name, x=x, y=y,
        marker=dict(
            color=color,
            opacity=opacity,
            line=dict(color="rgba(0,0,0,0.25)", width=1),
            # Gradiente: más claro arriba, más oscuro abajo
            pattern=dict(shape=""),  # sin pattern, solo color sólido con borde
        ),
        text=text,
        textposition=pos,
        textfont=dict(color=BLANCO, size=12, family="Inter"),
    )


def chart(fig):
    """Renderiza un plotly chart."""
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def img_to_b64(path: str) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()


def metric_card(label: str, value, suffix: str = "") -> str:
    return f"""
    <div style="background:{AZUL_MEDIO};border-radius:10px;padding:16px 20px;
                border-left:4px solid {DORADO};margin-bottom:8px;height:90px;">
        <div style="color:{GRIS_MEDIO};font-size:0.7rem;text-transform:uppercase;
                    letter-spacing:0.08em;margin-bottom:6px;">{label}</div>
        <div style="color:{BLANCO};font-size:2rem;font-weight:700;line-height:1;">
            {value}<span style="font-size:1rem;color:{AZUL_CELESTE};margin-left:2px">{suffix}</span>
        </div>
    </div>"""


def section(title: str):
    st.markdown(
        f"<h4 style='color:{DORADO};margin:24px 0 10px 0;font-size:1rem;"
        f"text-transform:uppercase;letter-spacing:0.06em;'>{title}</h4>",
        unsafe_allow_html=True,
    )


# ── Sidebar ───────────────────────────────────────────────────────────────────
NAV_ITEMS = [
    ("🏠", "Inicio"),
    ("📋", "Temporada"),
    ("🎮", "Partido"),
    ("⚔️", "Rivales"),
    ("💪", "Física"),
]

def render_sidebar():
    if "page" not in st.session_state:
        st.session_state.page = "Inicio"

    logo_arg = img_to_b64(os.path.join(ASSETS, "logo_argentino.png"))
    logo_sdc = img_to_b64(os.path.join(ASSETS, "sport data campus.png"))

    st.sidebar.markdown(f"""
    <div style="display:flex;align-items:center;justify-content:center;
                gap:14px;padding:20px 8px 16px 8px;">
        <img src="data:image/png;base64,{logo_arg}"
             style="width:68px;height:68px;object-fit:contain;border-radius:50%;
                    border:2px solid {DORADO};">
        <img src="data:image/png;base64,{logo_sdc}"
             style="width:68px;height:68px;object-fit:contain;border-radius:8px;">
    </div>
    <div style="text-align:center;color:{AZUL_CELESTE};font-size:0.78rem;
                font-weight:600;letter-spacing:0.06em;margin-bottom:16px;">
        ANÁLISIS DE RENDIMIENTO<br>
        <span style="color:{DORADO};font-size:0.85rem;">Temporada 2025–2026</span>
    </div>
    """, unsafe_allow_html=True)

    # Cajitas de navegación
    for icon, label in NAV_ITEMS:
        active = st.session_state.page == label
        # Wrapper div con clase activa/inactiva para que el CSS la pille
        st.sidebar.markdown(
            f'<div class="nav-item-{"active" if active else "inactive"}">',
            unsafe_allow_html=True,
        )
        if st.sidebar.button(f"{icon}  {label}", key=f"nav_{label}",
                             use_container_width=True):
            st.session_state.page = label
            st.rerun()
        st.sidebar.markdown("</div>", unsafe_allow_html=True)

    st.sidebar.markdown(f"<hr style='border-color:{AZUL_CLARO};margin:20px 0 12px'>",
                        unsafe_allow_html=True)
    st.sidebar.markdown(f"""
    <div style="text-align:center;color:{GRIS_MEDIO};font-size:0.67rem;
                line-height:1.7;padding-bottom:8px;">
        Desarrollado por<br>
        <span style="color:{DORADO};font-weight:700;font-size:0.78rem;">Ramón Codesido</span><br>
        <span style="color:{AZUL_CELESTE};font-size:0.72rem;">Sport Data Campus</span>
    </div>
    """, unsafe_allow_html=True)

    return st.session_state.page


# ── Página: Inicio ────────────────────────────────────────────────────────────
def page_inicio(df: pd.DataFrame):
    st.markdown(f"""
    <h1 style="color:{BLANCO};font-size:2rem;font-weight:700;margin-bottom:2px;">
        Club Argentino
    </h1>
    <p style="color:{AZUL_CELESTE};font-size:0.9rem;margin-bottom:20px;">
        Panel de análisis de rendimiento · Temporada 2025–2026
    </p>""", unsafe_allow_html=True)

    temporada = df[df["nivel"] == "temporada"]
    n        = len(temporada)
    goles_f  = int(temporada["goles"].fillna(0).sum())
    goles_c  = int(temporada["goles_enc"].fillna(0).sum())
    goles_c_str = str(goles_c)
    prom_pos = temporada["pct_posesion"].mean()
    prom_pas = temporada["pct_pases"].mean()
    victorias = sum(1 for _, r in temporada.iterrows()
                    if (r.get("goles") or 0) > (r.get("goles_enc") or 0))
    empates   = sum(1 for _, r in temporada.iterrows()
                    if (r.get("goles") or 0) == (r.get("goles_enc") or 0))
    derrotas  = n - victorias - empates

    # ── KPIs ──
    cols = st.columns(4)
    for col, (lbl, val, suf) in zip(cols, [
        ("Partidos jugados", n,                  ""),
        ("Posesión media",   f"{prom_pos:.1f}",  "%"),
        ("Precisión pases",  f"{prom_pas:.1f}",  "%"),
        ("Goles marcados",   goles_f,             ""),
    ]):
        col.markdown(metric_card(lbl, val, suf), unsafe_allow_html=True)

    # ── Balance V/E/D ──
    st.markdown(f"""
    <div style="display:flex;gap:10px;margin:12px 0 24px;">
        <div style="flex:1;background:{AZUL_MEDIO};border-radius:8px;padding:10px 0;
                    text-align:center;border-top:3px solid {VERDE};">
            <div style="color:{GRIS_MEDIO};font-size:0.65rem;text-transform:uppercase;
                        letter-spacing:0.08em;">Victorias</div>
            <div style="color:{VERDE};font-size:1.6rem;font-weight:800;">{victorias}</div>
        </div>
        <div style="flex:1;background:{AZUL_MEDIO};border-radius:8px;padding:10px 0;
                    text-align:center;border-top:3px solid {DORADO};">
            <div style="color:{GRIS_MEDIO};font-size:0.65rem;text-transform:uppercase;
                        letter-spacing:0.08em;">Empates</div>
            <div style="color:{DORADO};font-size:1.6rem;font-weight:800;">{empates}</div>
        </div>
        <div style="flex:1;background:{AZUL_MEDIO};border-radius:8px;padding:10px 0;
                    text-align:center;border-top:3px solid {ROJO};">
            <div style="color:{GRIS_MEDIO};font-size:0.65rem;text-transform:uppercase;
                        letter-spacing:0.08em;">Derrotas</div>
            <div style="color:{ROJO};font-size:1.6rem;font-weight:800;">{derrotas}</div>
        </div>
        <div style="flex:1;background:{AZUL_MEDIO};border-radius:8px;padding:10px 0;
                    text-align:center;border-top:3px solid {AZUL_CELESTE};">
            <div style="color:{GRIS_MEDIO};font-size:0.65rem;text-transform:uppercase;
                        letter-spacing:0.08em;">GF / GC</div>
            <div style="color:{BLANCO};font-size:1.6rem;font-weight:800;">
                <span style="color:{VERDE}">{goles_f}</span>
                <span style="color:{GRIS_MEDIO};font-size:1rem;"> / </span>
                <span style="color:{ROJO}">{goles_c_str}</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Resultados ──
    section("Resultados de la temporada")
    cols_h = ["Partido", "GF", "GC", "DIF", "Posesión", "Pases %"]
    st.markdown(f"""
    <div style="display:grid;grid-template-columns:2fr 0.7fr 0.7fr 0.7fr 1fr 1fr;
                gap:4px;margin-bottom:6px;padding:0 14px;">
        {"".join(f'<div style="color:{GRIS_MEDIO};font-size:0.65rem;text-transform:uppercase;'
                 f'letter-spacing:0.08em;font-weight:600;text-align:center;">{h}</div>'
                 for h in cols_h)}
    </div>""", unsafe_allow_html=True)

    for _, row in temporada.iterrows():
        gf    = int(row.get("goles") or 0)
        gc    = int(row.get("goles_enc") or 0)
        dif   = gf - gc
        pos   = row.get("pct_posesion") or 0
        pas   = row.get("pct_pases") or 0
        rival = row.get("rival", "")
        dif_str   = f"+{dif}" if dif > 0 else str(dif)
        res_color = VERDE if gf > gc else (ROJO if gf < gc else DORADO)
        dif_color = VERDE if dif > 0 else (ROJO if dif < 0 else DORADO)
        st.markdown(f"""
        <div style="display:grid;grid-template-columns:2fr 0.7fr 0.7fr 0.7fr 1fr 1fr;
                    gap:4px;background:{AZUL_MEDIO};border-radius:8px;
                    padding:11px 14px;margin-bottom:5px;
                    border-left:4px solid {res_color};align-items:center;">
            <div style="color:{BLANCO};font-weight:600;">CAdF vs {rival}</div>
            <div style="color:{VERDE};font-size:1.2rem;font-weight:700;text-align:center;">{gf}</div>
            <div style="color:{ROJO};font-size:1.2rem;font-weight:700;text-align:center;">{gc}</div>
            <div style="color:{dif_color};font-weight:700;text-align:center;">{dif_str}</div>
            <div style="color:{AZUL_CELESTE};font-weight:600;text-align:center;">{pos:.1f}%</div>
            <div style="color:{DORADO};font-weight:600;text-align:center;">{pas:.0f}%</div>
        </div>""", unsafe_allow_html=True)

    # ── Descripción de secciones ──
    section("Qué encontrarás en esta app")
    for icono, titulo, desc in [
        ("📋", "Temporada",
         "Evolución de métricas partido a partido: posesión, precisión de pases, tiros y goles. "
         "Heatmap de rendimiento por tramos de 15 minutos a lo largo de la temporada."),
        ("🎮", "Partido",
         "Análisis detallado de cada partido: línea de vida del control del juego por tramo, "
         "comparativa de mitades, campograma de pases por tercio del campo."),
        ("⚔️", "Rivales",
         "Histórico de enfrentamientos contra cada rival. Detecta en qué tramos "
         "pierdes el control y prepara tácticamente los próximos duelos."),
    ]:
        st.markdown(f"""
        <div style="background:{AZUL_MEDIO};border-radius:8px;padding:12px 16px;
                    margin-bottom:6px;display:flex;gap:14px;align-items:flex-start;">
            <div style="font-size:1.4rem;margin-top:2px;">{icono}</div>
            <div>
                <div style="color:{DORADO};font-weight:700;font-size:0.88rem;
                            margin-bottom:3px;">{titulo}</div>
                <div style="color:{GRIS_MEDIO};font-size:0.82rem;line-height:1.5;">{desc}</div>
            </div>
        </div>""", unsafe_allow_html=True)


# ── Página: Temporada ─────────────────────────────────────────────────────────
def _cell_color(val):
    if val >= 60: return "rgba(106,175,230,0.9)", AZUL_OSCURO
    if val >= 50: return "rgba(106,175,230,0.5)", BLANCO
    if val >= 45: return "rgba(201,168,76,0.55)", BLANCO
    return "rgba(232,93,117,0.75)", BLANCO


def _timeline_chart(temporada_df):
    """Gráfico combinado diferencia por partido + acumulado temporada."""
    p = temporada_df[["rival","goles","goles_enc"]].copy()
    p["gf"]  = p["goles"].fillna(0).astype(int)
    p["gc"]  = p["goles_enc"].fillna(0).astype(int)
    p["dif"] = p["gf"] - p["gc"]
    p["acum_dif"] = p["dif"].cumsum()
    etiquetas  = [f"vs {r}  {gf}–{gc}" for r, gf, gc in zip(p["rival"], p["gf"], p["gc"])]
    colors_dif = [VERDE if d > 0 else (ROJO if d < 0 else DORADO) for d in p["dif"]]
    acum = p["acum_dif"].tolist()
    partidos_raw = p

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Bar(
        x=etiquetas, y=partidos_raw["dif"], name="Diferencia por partido",
        marker=dict(color=colors_dif, line=dict(color=AZUL_OSCURO, width=1), opacity=0.85),
        text=[f"+{d}" if d > 0 else str(d) for d in partidos_raw["dif"]],
        textposition="outside", textfont=dict(color=BLANCO, size=13),
    ), secondary_y=False)
    fig.add_trace(go.Scatter(
        x=etiquetas, y=acum, name="Acumulado temporada",
        mode="lines+markers+text",
        line=dict(color=DORADO, width=3, shape="spline"),
        marker=dict(size=12, color=[VERDE if v > 0 else (ROJO if v < 0 else DORADO) for v in acum],
                    line=dict(color=BLANCO, width=2)),
        text=[f"+{v}" if v > 0 else str(v) for v in acum],
        textposition="top center", textfont=dict(color=DORADO, size=11),
    ), secondary_y=True)
    fig.add_hline(y=0, line_color="rgba(255,255,255,0.2)", line_width=1.5,
                  line_dash="dot", secondary_y=False)
    fig.update_layout(
        paper_bgcolor=AZUL_OSCURO, plot_bgcolor=PLOT_BG,
        height=340, margin=dict(t=30, b=20, l=40, r=60),
        font=dict(family="Inter, Arial, sans-serif", color=BLANCO),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=BLANCO),
                    orientation="h", y=1.12),
        bargap=0.35,
    )
    fig.update_xaxes(gridcolor="rgba(200,214,229,0.08)", tickfont=dict(color=BLANCO, size=11))
    fig.update_yaxes(title_text="Diferencia", gridcolor="rgba(200,214,229,0.08)",
                     tickfont=dict(color=BLANCO), secondary_y=False,
                     zeroline=True, zerolinecolor="rgba(255,255,255,0.15)")
    fig.update_yaxes(title_text="Acumulado", gridcolor="rgba(0,0,0,0)",
                     tickfont=dict(color=DORADO), color=DORADO, secondary_y=True)
    return fig


def page_temporada(df: pd.DataFrame):
    st.markdown(f"<h2 style='color:{BLANCO};margin-bottom:4px;'>📋 Temporada</h2>",
                unsafe_allow_html=True)

    temporada = df[df["nivel"] == "temporada"].copy()
    rivales   = temporada["rival"].tolist()
    n_partidos = max(len(rivales), 1)

    # ── 1: Heatmap posesión por tramo ──
    section("Mapa de posesión por tramo y partido")
    tramos_df = df[df["nivel"] == "tramo"].copy()
    pivot = (
        tramos_df.pivot_table(index="rival", columns="tramo_raw", values="pct_posesion")
        .reindex(columns=TRAMO_ORDER)
    )
    promedios = pivot.mean()
    n_rivales = len(pivot.index)
    n_tramos  = len(TRAMO_ORDER)
    prom_y    = n_rivales + 0.3

    fig_hm = go.Figure()
    for r_idx, rival in enumerate(pivot.index):
        for t_idx, tramo in enumerate(TRAMO_ORDER):
            val = pivot.loc[rival, tramo] if tramo in pivot.columns else None
            if pd.isna(val): continue
            bg, tc = _cell_color(val)
            fig_hm.add_shape(type="rect",
                x0=t_idx-0.45, x1=t_idx+0.45, y0=r_idx-0.4, y1=r_idx+0.4,
                fillcolor=bg, line=dict(color="rgba(255,255,255,0.06)", width=1))
            fig_hm.add_annotation(x=t_idx, y=r_idx, text=f"<b>{val:.0f}%</b>",
                showarrow=False, font=dict(color=tc, size=13, family="Inter"))
    for t_idx, tramo in enumerate(TRAMO_ORDER):
        pval = promedios.get(tramo)
        if pd.isna(pval): continue
        bg, tc = _cell_color(pval)
        fig_hm.add_shape(type="rect",
            x0=t_idx-0.45, x1=t_idx+0.45,
            y0=prom_y-0.38, y1=prom_y+0.38,
            fillcolor=bg, line=dict(color=DORADO, width=1.5))
        fig_hm.add_annotation(x=t_idx, y=prom_y, text=f"<b>{pval:.0f}%</b>",
            showarrow=False, font=dict(color=tc, size=12, family="Inter"))

    fig_hm.update_layout(
        paper_bgcolor=AZUL_OSCURO, plot_bgcolor=AZUL_OSCURO,
        height=max(220, (n_rivales + 1) * 85 + 80),
        margin=dict(t=20, b=50, l=90, r=20),
        xaxis=dict(tickvals=list(range(n_tramos)), ticktext=TRAMO_ORDER,
                   showgrid=False, zeroline=False, tickfont=dict(color=GRIS_MEDIO, size=11)),
        yaxis=dict(tickvals=list(range(n_rivales)) + [prom_y],
                   ticktext=list(pivot.index) + ["<b>Promedio</b>"],
                   showgrid=False, zeroline=False, tickfont=dict(color=BLANCO, size=11),
                   autorange="reversed"),
    )
    for color, label in [("rgba(106,175,230,0.9)", "≥60%"),("rgba(106,175,230,0.5)", "50–60%"),
                          ("rgba(201,168,76,0.55)", "45–50%"),("rgba(232,93,117,0.75)", "<45%")]:
        fig_hm.add_trace(go.Scatter(x=[None], y=[None], mode="markers",
            marker=dict(size=11, color=color, symbol="square"), name=label, showlegend=True))
    fig_hm.update_layout(legend=dict(orientation="h", y=-0.18, x=0,
        font=dict(color=GRIS_MEDIO, size=10), bgcolor="rgba(0,0,0,0)"))
    chart(fig_hm)

    # ── 2: Promedio 1ª vs 2ª mitad ──
    section("Promedio 1ª vs 2ª Mitad — toda la temporada")
    partes_df = df[df["nivel"] == "parte"]
    p1m = partes_df[partes_df["tramo_raw"] == "1º mitad"].mean(numeric_only=True)
    p2m = partes_df[partes_df["tramo_raw"] == "2º mitad"].mean(numeric_only=True)
    mets = {"Posesión %": "pct_posesion", "Pases %": "pct_pases",
            "Tiros": "tiros", "Centros": "centros", "Tiros puerta": "tiros_puerta"}
    v1 = [p1m.get(v, 0) or 0 for v in mets.values()]
    v2 = [p2m.get(v, 0) or 0 for v in mets.values()]
    fig_mit = go.Figure()
    fig_mit.add_trace(go.Bar(name="1ª Mitad (prom)", x=list(mets.keys()), y=v1,
        marker_color=AZUL_CELESTE, text=[f"{v:.1f}" for v in v1],
        textposition="outside", textfont=dict(color=BLANCO)))
    fig_mit.add_trace(go.Bar(name="2ª Mitad (prom)", x=list(mets.keys()), y=v2,
        marker_color=DORADO, text=[f"{v:.1f}" for v in v2],
        textposition="outside", textfont=dict(color=BLANCO)))
    t(fig_mit, height=320, barmode="group", margin=dict(t=20, b=20, l=20, r=20))
    chart(fig_mit)

    # ── 3: Posesión + Pases % ──
    section("Control del juego — Posesión & Precisión de pases")
    fig_ctrl = go.Figure()
    fig_ctrl.add_trace(go.Scatter(
        x=rivales, y=temporada["pct_posesion"], name="Posesión %",
        mode="lines+markers",
        line=dict(color=AZUL_CELESTE, width=3),
        marker=dict(size=11, color=AZUL_CELESTE, line=dict(color=BLANCO, width=2)),
        fill="tozeroy", fillcolor="rgba(106,175,230,0.08)",
    ))
    fig_ctrl.add_trace(go.Scatter(
        x=rivales, y=temporada["pct_pases"], name="Precisión pases %",
        mode="lines+markers",
        line=dict(color=DORADO, width=3, dash="dot"),
        marker=dict(size=11, color=DORADO, line=dict(color=BLANCO, width=2)),
    ))
    fig_ctrl.add_hline(y=50, line_dash="dash", line_color="rgba(255,255,255,0.2)", line_width=1)
    for rv, pos, pas in zip(rivales, temporada["pct_posesion"], temporada["pct_pases"]):
        fig_ctrl.add_annotation(x=rv, y=pos, text=f"{pos:.1f}%", showarrow=False,
                                yshift=16, font=dict(color=AZUL_CELESTE, size=10))
        fig_ctrl.add_annotation(x=rv, y=pas, text=f"{pas:.0f}%", showarrow=False,
                                yshift=-18, font=dict(color=DORADO, size=10))
    t(fig_ctrl, height=360, yaxis=dict(range=[30, 105], ticksuffix="%",
                                       gridcolor="rgba(200,214,229,0.1)",
                                       tickfont=dict(color=BLANCO)))
    chart(fig_ctrl)

    # ── 4: Pases por tercio — campograma con selector ──
    section("Pases por tercio del campo")
    modo = st.radio("", ["Promedio por partido", "Total temporada"],
                    horizontal=True, label_visibility="collapsed")

    tercios_data = [
        ("Zona defensiva\n(1er tercio)",   "def_intentados",  "def_completados",  "pct_pases_def"),
        ("Zona media\n(mitad campo)",       "mid_intentados",  "mid_completados",  "pct_pases_mid"),
        ("Zona de ataque\n(último tercio)", "atq_intentados",  "atq_completados",  "pct_pases_atq"),
    ]

    fig_pitch = go.Figure()
    pitch_w, pitch_h = 105, 68

    # Campo
    fig_pitch.add_shape(type="rect", x0=0, y0=0, x1=pitch_w, y1=pitch_h,
                        fillcolor="#2D5A1B", line=dict(color="white", width=2))
    fig_pitch.add_shape(type="line", x0=pitch_w/2, y0=0, x1=pitch_w/2, y1=pitch_h,
                        line=dict(color="white", width=1.5, dash="dot"))
    fig_pitch.add_shape(type="circle",
                        x0=pitch_w/2-9.15, y0=pitch_h/2-9.15,
                        x1=pitch_w/2+9.15, y1=pitch_h/2+9.15,
                        line=dict(color="white", width=1.5))
    for x0, x1 in [(0, 16.5), (pitch_w-16.5, pitch_w)]:
        fig_pitch.add_shape(type="rect", x0=x0, y0=pitch_h/2-20.16,
                            x1=x1, y1=pitch_h/2+20.16,
                            fillcolor="rgba(0,0,0,0)", line=dict(color="white", width=1.5))

    zone_colors = ["rgba(232,93,117,{a})", "rgba(201,168,76,{a})", "rgba(106,175,230,{a})"]
    zone_w = pitch_w / 3

    for i, (zona, ip_col, pc_col, pct_col) in enumerate(tercios_data):
        total_int  = temporada[ip_col].fillna(0).sum()
        total_comp = temporada[pc_col].fillna(0).sum()
        total_perd = total_int - total_comp
        # % real por partido (media de los % individuales, no calculado del total)
        avg_pct = temporada[pct_col].fillna(0).mean()

        if modo == "Promedio por partido":
            vi = total_int / n_partidos
            vc = total_comp / n_partidos
            vp = total_perd / n_partidos
            lbl_c, lbl_p, lbl_i = f"{vc:.1f}", f"{vp:.1f}", f"{vi:.1f}"
        else:
            vi, vc, vp = total_int, total_comp, total_perd
            lbl_c, lbl_p, lbl_i = str(int(vc)), str(int(vp)), str(int(vi))

        x0 = i * zone_w
        x1 = (i + 1) * zone_w
        cx = (x0 + x1) / 2
        alpha = 0.15 + (avg_pct / 100) * 0.35
        fig_pitch.add_shape(type="rect", x0=x0+1, y0=1, x1=x1-1, y1=pitch_h-1,
                            fillcolor=zone_colors[i].format(a=round(alpha, 2)),
                            line=dict(color="rgba(255,255,255,0.12)", width=1))

        pct_color = VERDE if avg_pct >= 80 else (DORADO if avg_pct >= 70 else ROJO)
        r = 9
        fig_pitch.add_shape(type="circle",
                            x0=cx-r, y0=pitch_h/2-r, x1=cx+r, y1=pitch_h/2+r,
                            fillcolor=AZUL_OSCURO, line=dict(color=pct_color, width=2.5))
        fig_pitch.add_annotation(x=cx, y=pitch_h/2, text=f"<b>{avg_pct:.0f}%</b>",
                                  showarrow=False,
                                  font=dict(color=pct_color, size=14, family="Inter"))
        fig_pitch.add_annotation(x=cx, y=pitch_h-5,
                                  text=f"<b>{zona.replace(chr(10),'<br>')}</b>",
                                  showarrow=False,
                                  font=dict(color="white", size=9, family="Inter"),
                                  align="center")
        fig_pitch.add_annotation(x=cx, y=8,
                                  text=f"✅ {lbl_c}   ❌ {lbl_p}",
                                  showarrow=False,
                                  font=dict(color=BLANCO, size=11, family="Inter"))

    fig_pitch.update_layout(
        paper_bgcolor=AZUL_OSCURO, plot_bgcolor=AZUL_OSCURO,
        height=360, margin=dict(t=10, b=10, l=10, r=10),
        xaxis=dict(range=[0, pitch_w], showgrid=False, zeroline=False,
                   showticklabels=False, scaleanchor="y", scaleratio=1),
        yaxis=dict(range=[0, pitch_h], showgrid=False, zeroline=False,
                   showticklabels=False),
        showlegend=False,
    )
    chart(fig_pitch)

    # ── 5: Tiros totales, a puerta, % eficacia ──
    section("Ataque — Tiros totales, a puerta y eficacia")
    fig_tir = make_subplots(specs=[[{"secondary_y": True}]])
    tiros_tot = temporada["tiros"].fillna(0).astype(int).tolist()
    tiros_prt = temporada["tiros_puerta"].fillna(0).astype(int).tolist()
    pct_efic  = [round(tp/tt*100, 1) if tt > 0 else 0
                 for tt, tp in zip(tiros_tot, tiros_prt)]
    fig_tir.add_trace(go.Bar(
        name="Tiros totales", x=rivales, y=tiros_tot,
        marker=dict(color=AZUL_CELESTE, opacity=0.7, line=dict(color=AZUL_OSCURO, width=1)),
        text=tiros_tot, textposition="outside", textfont=dict(color=AZUL_CELESTE, size=12),
    ), secondary_y=False)
    fig_tir.add_trace(go.Bar(
        name="Tiros a puerta", x=rivales, y=tiros_prt,
        marker=dict(color=DORADO, opacity=0.9, line=dict(color=AZUL_OSCURO, width=1)),
        text=tiros_prt, textposition="inside", textfont=dict(color=BLANCO, size=12),
    ), secondary_y=False)
    fig_tir.add_trace(go.Scatter(
        name="% Eficacia (tiros→puerta)", x=rivales, y=pct_efic,
        mode="lines+markers+text",
        line=dict(color=VERDE, width=3),
        marker=dict(size=12, color=VERDE, line=dict(color=BLANCO, width=2)),
        text=[f"{v}%" for v in pct_efic], textposition="top center",
        textfont=dict(color=VERDE, size=11),
    ), secondary_y=True)
    fig_tir.update_layout(
        paper_bgcolor=AZUL_OSCURO, plot_bgcolor=PLOT_BG,
        font=dict(family="Inter, Arial, sans-serif", color=BLANCO),
        height=380, margin=dict(t=30, b=30, l=40, r=70),
        barmode="group", bargap=0.25, bargroupgap=0.05,
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=BLANCO), orientation="h", y=1.1),
    )
    fig_tir.update_yaxes(title_text="Nº tiros", secondary_y=False,
                          gridcolor="rgba(200,214,229,0.1)", tickfont=dict(color=BLANCO))
    fig_tir.update_yaxes(title_text="% Eficacia", secondary_y=True,
                          gridcolor="rgba(0,0,0,0)", tickfont=dict(color=VERDE),
                          color=VERDE, range=[0, 100], ticksuffix="%")
    fig_tir.update_xaxes(gridcolor="rgba(200,214,229,0.1)", tickfont=dict(color=BLANCO))
    chart(fig_tir)

    # ── 6: Defensa — paradas vs goles encajados + faltas ──
    section("Defensa & Portero — evolución de temporada")

    par_t  = temporada["paradas"].fillna(0).astype(int).tolist()
    tap_t  = temporada["tap"].fillna(0).astype(int).tolist()
    ge_t   = temporada["goles_enc"].fillna(0).astype(int).tolist()
    fal_t  = temporada["faltas"].fillna(0).astype(int).tolist()
    sch_t  = temporada["jugadas_set"].fillna(0).astype(int).tolist()

    fig_def_t = make_subplots(specs=[[{"secondary_y": True}]])
    fig_def_t.add_trace(go.Bar(
        name="Paradas portero", x=rivales, y=par_t,
        marker=dict(color=AZUL_CELESTE, line=dict(color=AZUL_OSCURO, width=1)),
        text=par_t, textposition="outside",
        textfont=dict(color=AZUL_CELESTE, size=11),
    ), secondary_y=False)
    fig_def_t.add_trace(go.Bar(
        name="Disparos bloqueados", x=rivales, y=tap_t,
        marker=dict(color=DORADO, opacity=0.85, line=dict(color=AZUL_OSCURO, width=1)),
        text=tap_t, textposition="outside",
        textfont=dict(color=DORADO, size=11),
    ), secondary_y=False)
    fig_def_t.add_trace(go.Scatter(
        name="Goles encajados", x=rivales, y=ge_t,
        mode="lines+markers+text",
        line=dict(color=ROJO, width=2.5),
        marker=dict(size=10, color=ROJO, line=dict(color=BLANCO, width=1.5)),
        text=ge_t, textposition="top center",
        textfont=dict(color=ROJO, size=11),
    ), secondary_y=True)
    fig_def_t.update_layout(
        paper_bgcolor=PAPER_BG, plot_bgcolor=PLOT_BG,
        height=320, margin=dict(t=30, b=20, l=40, r=60),
        barmode="group", bargap=0.25,
        font=dict(family="Inter, Arial, sans-serif", color=BLANCO),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=BLANCO),
                    orientation="h", y=1.1),
    )
    fig_def_t.update_yaxes(title_text="Paradas / TAP", secondary_y=False,
                            gridcolor="rgba(200,214,229,0.07)",
                            tickfont=dict(color=BLANCO))
    fig_def_t.update_yaxes(title_text="Goles encajados", secondary_y=True,
                            gridcolor="rgba(0,0,0,0)",
                            tickfont=dict(color=ROJO), color=ROJO)
    fig_def_t.update_xaxes(gridcolor="rgba(200,214,229,0.07)",
                             tickfont=dict(color=BLANCO))
    chart(fig_def_t)

    # Jugadas ensayadas + faltas por partido
    section("Jugadas ensayadas y faltas — evolución de temporada")
    fig_je = go.Figure()
    fig_je.add_trace(go.Bar(
        name="Saques de esquina", x=rivales, y=sch_t,
        marker=dict(color=AZUL_CELESTE, line=dict(color=AZUL_OSCURO, width=1)),
        text=sch_t, textposition="outside",
        textfont=dict(color=AZUL_CELESTE, size=11),
    ))
    fig_je.add_trace(go.Bar(
        name="Faltas cometidas", x=rivales, y=fal_t,
        marker=dict(color=ROJO, opacity=0.8, line=dict(color=AZUL_OSCURO, width=1)),
        text=fal_t, textposition="outside",
        textfont=dict(color=ROJO, size=11),
    ))
    t(fig_je, height=280, barmode="group", bargap=0.3,
      margin=dict(t=20, b=20, l=30, r=20))
    chart(fig_je)

    # ── 7: Diferencia de goles y acumulado ──
    section("Diferencia de goles y acumulado de temporada")
    chart(_timeline_chart(temporada))


# ── Página: Partido ───────────────────────────────────────────────────────────
def page_partido(df: pd.DataFrame):
    st.markdown(f"<h2 style='color:{BLANCO};margin-bottom:4px;'>🎮 Partido</h2>",
                unsafe_allow_html=True)

    rivales   = df[df["nivel"] == "temporada"]["rival"].unique().tolist()
    rival_sel = st.selectbox("", rivales, format_func=lambda r: f"CAdF vs {r}",
                             label_visibility="collapsed")

    pdata    = df[df["rival"] == rival_sel]
    temp_row = pdata[pdata["nivel"] == "temporada"].iloc[0]
    tramos   = pdata[pdata["nivel"] == "tramo"].set_index("tramo_raw").reindex(TRAMO_ORDER).reset_index()
    partes   = pdata[pdata["nivel"] == "parte"]

    gf        = int(temp_row.get("goles") or 0)
    gc        = int(temp_row.get("goles_enc") or 0)
    gc_str    = str(gc)
    res_color = VERDE if gf > gc else (ROJO if gf < gc else DORADO)
    res_text  = "Victoria" if gf > gc else ("Derrota" if gf < gc else "Empate")

    # ── Banner resultado ──
    st.markdown(f"""
    <div style="background:{AZUL_MEDIO};border-radius:12px;padding:20px 28px;
                margin-bottom:20px;border-left:6px solid {res_color};
                display:flex;align-items:center;gap:32px;">
        <div>
            <div style="color:{GRIS_MEDIO};font-size:0.7rem;text-transform:uppercase;
                        letter-spacing:0.1em;">Resultado</div>
            <div style="color:{BLANCO};font-size:3rem;font-weight:800;line-height:1.1;">
                {gf} <span style="color:{GRIS_MEDIO};font-size:1.5rem;">–</span> {gc_str}
            </div>
            <div style="color:{res_color};font-size:0.85rem;font-weight:600;">{res_text}</div>
        </div>
        <div style="display:flex;gap:20px;">
            <div style="text-align:center;">
                <div style="color:{GRIS_MEDIO};font-size:0.65rem;text-transform:uppercase;">Posesión</div>
                <div style="color:{AZUL_CELESTE};font-size:1.5rem;font-weight:700;">
                    {temp_row.get('pct_posesion') or 0:.1f}%</div>
            </div>
            <div style="text-align:center;">
                <div style="color:{GRIS_MEDIO};font-size:0.65rem;text-transform:uppercase;">Pases %</div>
                <div style="color:{DORADO};font-size:1.5rem;font-weight:700;">
                    {temp_row.get('pct_pases') or 0:.0f}%</div>
            </div>
            <div style="text-align:center;">
                <div style="color:{GRIS_MEDIO};font-size:0.65rem;text-transform:uppercase;">Tiros puerta</div>
                <div style="color:{BLANCO};font-size:1.5rem;font-weight:700;">
                    {int(temp_row.get('tiros_puerta') or 0)}</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Match Momentum ──
    section("Match Momentum — Posesión por tramo")

    tramos_mm   = tramos.dropna(subset=["pct_posesion"])
    tramo_labels_mm = tramos_mm["tramo_raw"].tolist()
    pos_vals_mm = tramos_mm["pct_posesion"].tolist()
    goles_mm    = tramos_mm["goles"].tolist()

    # Valor CAdF: pos - 50 (>0 domina CAdF, <0 domina rival)
    diff_vals   = [p - 50 for p in pos_vals_mm]
    colors_mm   = [AZUL_CELESTE if d >= 0 else GRIS_MEDIO for d in diff_vals]

    fig_mm = go.Figure()

    # Zona de fondo: arriba CAdF, abajo rival
    fig_mm.add_hrect(y0=0,   y1=50,  fillcolor="rgba(106,175,230,0.05)", line_width=0)
    fig_mm.add_hrect(y0=-50, y1=0,   fillcolor="rgba(200,214,229,0.04)", line_width=0)
    fig_mm.add_hline(y=0, line_color="rgba(255,255,255,0.35)", line_width=1.5)

    # Barras CAdF (arriba) — usando Bar trace real
    fig_mm.add_trace(go.Bar(
        x=tramo_labels_mm,
        y=[max(d, 0) for d in diff_vals],
        marker=dict(
            color=[AZUL_CELESTE if d >= 0 else "rgba(0,0,0,0)" for d in diff_vals],
            line=dict(width=0),
        ),
        base=0,
        showlegend=False,
        hovertemplate="%{x}<br>CAdF: %{customdata:.0f}%<extra></extra>",
        customdata=pos_vals_mm,
    ))
    # Barras rival (abajo) — invertidas
    fig_mm.add_trace(go.Bar(
        x=tramo_labels_mm,
        y=[min(d, 0) for d in diff_vals],
        marker=dict(
            color=[GRIS_MEDIO if d < 0 else "rgba(0,0,0,0)" for d in diff_vals],
            opacity=0.7,
            line=dict(width=0),
        ),
        base=0,
        showlegend=False,
        hovertemplate="%{x}<br>Rival: %{customdata:.0f}%<extra></extra>",
        customdata=[100 - p for p in pos_vals_mm],
    ))

    # Etiquetas % posesión CAdF
    for lbl, d, pos in zip(tramo_labels_mm, diff_vals, pos_vals_mm):
        yshift = 8 if d >= 0 else -8
        yanchor = "bottom" if d >= 0 else "top"
        fig_mm.add_annotation(
            x=lbl, y=max(d, 0) if d >= 0 else min(d, 0),
            text=f"<b>{pos:.0f}%</b>",
            showarrow=False, yshift=yshift,
            yanchor=yanchor,
            font=dict(color=AZUL_CELESTE if d >= 0 else GRIS_MEDIO,
                      size=10, family="Inter"),
        )

    # ⚽ goles
    for lbl, d, g in zip(tramo_labels_mm, diff_vals, goles_mm):
        if pd.notna(g) and g > 0:
            for k in range(int(g)):
                fig_mm.add_annotation(
                    x=lbl, y=max(d, 0) + 10 + k * 10,
                    text="⚽", showarrow=False, font=dict(size=15),
                )

    # Separador descanso
    if "45+" in tramo_labels_mm and "46-60" in tramo_labels_mm:
        fig_mm.add_vline(
            x=tramo_labels_mm.index("45+") + 0.5,
            line_color="rgba(255,255,255,0.25)", line_width=1.5, line_dash="dash",
        )
        fig_mm.add_annotation(
            x=tramo_labels_mm.index("45+") + 0.5, y=48,
            text="D", showarrow=False,
            font=dict(color=GRIS_MEDIO, size=10, family="Inter"),
        )

    # Etiquetas CAdF / Rival
    fig_mm.add_annotation(x=tramo_labels_mm[0], y=48, text=f"<b>CAdF</b>",
                          showarrow=False, xanchor="left",
                          font=dict(color=AZUL_CELESTE, size=11, family="Inter"))
    fig_mm.add_annotation(x=tramo_labels_mm[0], y=-48, text=f"<b>{rival_sel}</b>",
                          showarrow=False, xanchor="left",
                          font=dict(color=GRIS_MEDIO, size=11, family="Inter"))

    fig_mm.update_layout(
        paper_bgcolor=PAPER_BG, plot_bgcolor=PLOT_BG,
        height=320, margin=dict(t=20, b=30, l=60, r=20),
        font=dict(family="Inter, Arial, sans-serif", color=BLANCO),
        barmode="overlay", bargap=0.25,
        showlegend=False,
        xaxis=dict(showgrid=False, zeroline=False,
                   tickfont=dict(color=GRIS_MEDIO, size=11)),
        yaxis=dict(
            range=[-55, 60],
            tickvals=[-50, -25, 0, 25, 50],
            ticktext=["100%", "75%", "50%", "75%", "100%"],
            showgrid=False, zeroline=False,
            tickfont=dict(color=GRIS_MEDIO, size=10),
        ),
        shapes=[dict(type="rect", xref="paper", yref="paper",
                     x0=0, y0=0, x1=1, y1=1,
                     line=dict(color="rgba(106,175,230,0.12)", width=1),
                     fillcolor="rgba(0,0,0,0)", layer="above")],
    )
    chart(fig_mm)

    # ── Línea de vida del partido ──
    section("Línea de vida — control del partido por tramo")

    fig_vida = go.Figure()

    # Gradiente de fondo: zona de dominio
    fig_vida.add_hrect(y0=50, y1=100, fillcolor="rgba(106,175,230,0.04)", line_width=0)
    fig_vida.add_hrect(y0=0,  y1=50,  fillcolor="rgba(232,93,117,0.04)",  line_width=0)
    fig_vida.add_hline(y=50, line_dash="dot",
                       line_color="rgba(200,214,229,0.35)", line_width=1.5,
                       annotation_text="50%", annotation_font=dict(color=GRIS_MEDIO, size=9))

    pos_vals  = tramos["pct_posesion"].tolist()
    pas_vals  = tramos["pct_pases"].tolist()
    tramo_labels = tramos["tramo_raw"].tolist()

    # Área de posesión con gradiente
    fig_vida.add_trace(go.Scatter(
        x=tramo_labels, y=pos_vals,
        name="Posesión %", mode="lines+markers",
        line=dict(color=AZUL_CELESTE, width=3.5, shape="spline"),
        marker=dict(size=12, color=AZUL_CELESTE,
                    line=dict(color=BLANCO, width=2.5)),
        fill="tozeroy",
        fillcolor="rgba(106,175,230,0.15)",
        hovertemplate="Tramo %{x}<br>Posesión: <b>%{y:.1f}%</b><extra></extra>",
    ))
    fig_vida.add_trace(go.Scatter(
        x=tramo_labels, y=pas_vals,
        name="Precisión pases %", mode="lines+markers",
        line=dict(color=DORADO, width=2.5, dash="dash", shape="spline"),
        marker=dict(size=9, color=DORADO),
        hovertemplate="Tramo %{x}<br>Pases: <b>%{y:.1f}%</b><extra></extra>",
    ))

    # Anotar goles a favor (⚽ arriba) y en contra (🔴 abajo) por tramo
    for _, trow in tramos.iterrows():
        if pd.notna(trow.get("goles")) and trow["goles"] > 0:
            for _ in range(int(trow["goles"])):
                fig_vida.add_annotation(
                    x=trow["tramo_raw"], y=96,
                    text="⚽", showarrow=False, font=dict(size=18),
                )
        if pd.notna(trow.get("goles_enc")) and trow["goles_enc"] > 0:
            for _ in range(int(trow["goles_enc"])):
                fig_vida.add_annotation(
                    x=trow["tramo_raw"], y=6,
                    text="🔴", showarrow=False, font=dict(size=15),
                )

    # Etiquetas de posesión por tramo
    for i, (tramo, val) in enumerate(zip(tramo_labels, pos_vals)):
        if pd.notna(val):
            color = AZUL_CELESTE if val >= 50 else ROJO
            fig_vida.add_annotation(
                x=tramo, y=val,
                text=f"{val:.0f}%", showarrow=False,
                yshift=18, font=dict(color=color, size=9, family="Inter"),
            )

    fig_vida.update_layout(
        paper_bgcolor=AZUL_OSCURO, plot_bgcolor=PLOT_BG,
        height=400, margin=dict(t=20, b=40, l=40, r=20),
        font=dict(family="Inter, Arial, sans-serif", color=BLANCO),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=BLANCO),
                    orientation="h", y=1.08),
        yaxis=dict(range=[0, 110], ticksuffix="%",
                   gridcolor="rgba(200,214,229,0.08)",
                   tickfont=dict(color=BLANCO)),
        xaxis=dict(gridcolor="rgba(200,214,229,0.08)",
                   tickfont=dict(color=BLANCO, size=11)),
    )
    chart(fig_vida)

    # ── Comparativa 1ª vs 2ª mitad ──
    section("1ª vs 2ª Mitad")
    p1 = partes[partes["tramo_raw"] == "1º mitad"].iloc[0] if len(partes) >= 1 else {}
    p2 = partes[partes["tramo_raw"] == "2º mitad"].iloc[0] if len(partes) >= 2 else {}

    mets = {"Posesión %": "pct_posesion", "Pases %": "pct_pases",
            "Tiros": "tiros", "Centros": "centros", "Tiros puerta": "tiros_puerta"}
    labels = list(mets.keys())
    v1 = [p1.get(v) or 0 for v in mets.values()]
    v2 = [p2.get(v) or 0 for v in mets.values()]

    fig_m = go.Figure()
    fig_m.add_trace(go.Bar(name="1ª Mitad", x=labels, y=v1,
                           marker_color=AZUL_CELESTE,
                           text=[f"{v:.0f}" for v in v1],
                           textposition="outside", textfont=dict(color=BLANCO)))
    fig_m.add_trace(go.Bar(name="2ª Mitad", x=labels, y=v2,
                           marker_color=DORADO,
                           text=[f"{v:.0f}" for v in v2],
                           textposition="outside", textfont=dict(color=BLANCO)))
    t(fig_m, height=320, barmode="group", margin=dict(t=20, b=20, l=20, r=20))
    chart(fig_m)

    # ── Campograma de pases por tercio ──
    section("Pases por tercio del campo")

    tercios_data = [
        ("DEF", "def_intentados", "def_completados", "pct_pases_def"),
        ("MED", "mid_intentados", "mid_completados", "pct_pases_mid"),
        ("ATQ", "atq_intentados", "atq_completados", "pct_pases_atq"),
    ]

    # Recoger datos
    zonas_vals = []
    for label, ip, pc, pct_col in tercios_data:
        intentados  = int(temp_row.get(ip) or 0)
        completados = int(temp_row.get(pc) or 0)
        perdidos    = intentados - completados
        pct         = temp_row.get(pct_col) or 0
        zonas_vals.append((label, intentados, completados, perdidos, pct))

    total_intentados = sum(v[1] for v in zonas_vals) or 1

    # ── Campograma + Barras: un único figura con subplots compartiendo eje X ──
    _, col_main, _ = st.columns([1, 7, 1])
    with col_main:
        pitch_w, pitch_h = 105, 68
        zone_w  = pitch_w / 3
        max_int = max(v[1] for v in zonas_vals) or 1

        # Centros de cada zona en coordenadas del campo
        zone_cx = [zone_w/2, zone_w + zone_w/2, 2*zone_w + zone_w/2]  # 17.5, 52.5, 87.5

        fig_cp = make_subplots(
            rows=2, cols=1,
            row_heights=[0.62, 0.38],
            vertical_spacing=0.03,
            shared_xaxes=True,
        )

        # ── Shapes del campo: layer="below" para que las burbujas queden encima ──
        def ps(shape_dict):
            """Añade layer=below a un shape del campo."""
            shape_dict["layer"] = "below"
            shape_dict.setdefault("xref", "x")
            shape_dict.setdefault("yref", "y")
            return shape_dict

        pitch_shapes = []
        # Césped
        for i in range(6):
            pitch_shapes.append(ps(dict(type="rect",
                x0=i*(pitch_w/6), y0=0, x1=(i+1)*(pitch_w/6), y1=pitch_h,
                fillcolor="#1E3D14" if i%2==0 else "#243D18",
                line=dict(width=0))))
        # Borde
        pitch_shapes.append(ps(dict(type="rect", x0=0, y0=0, x1=pitch_w, y1=pitch_h,
            fillcolor="rgba(0,0,0,0)", line=dict(color="white", width=2))))
        # Línea central
        pitch_shapes.append(ps(dict(type="line", x0=pitch_w/2, y0=0,
            x1=pitch_w/2, y1=pitch_h, line=dict(color="white", width=1.5))))
        # Círculo central
        pitch_shapes.append(ps(dict(type="circle",
            x0=pitch_w/2-9.15, y0=pitch_h/2-9.15,
            x1=pitch_w/2+9.15, y1=pitch_h/2+9.15,
            fillcolor="rgba(0,0,0,0)", line=dict(color="white", width=1.5))))
        # Áreas
        for ax0, ax1 in [(0, 16.5), (pitch_w-16.5, pitch_w)]:
            pitch_shapes.append(ps(dict(type="rect", x0=ax0, y0=pitch_h/2-20.16,
                x1=ax1, y1=pitch_h/2+20.16, fillcolor="rgba(0,0,0,0)",
                line=dict(color="white", width=1.5))))
            px0 = ax0 if ax0==0 else pitch_w-5.5
            px1 = 5.5 if ax0==0 else pitch_w
            pitch_shapes.append(ps(dict(type="rect", x0=px0, y0=pitch_h/2-9.16,
                x1=px1, y1=pitch_h/2+9.16, fillcolor="rgba(0,0,0,0)",
                line=dict(color="white", width=1))))
        # Divisores de tercio
        for xi in [zone_w, 2*zone_w]:
            pitch_shapes.append(ps(dict(type="line", x0=xi, y0=0, x1=xi, y1=pitch_h,
                line=dict(color="rgba(255,255,255,0.35)", width=1.5, dash="dash"))))

        # ── Row 1: burbujas (trace, siempre encima de shapes) ──
        bubble_sizes, bubble_texts = [], []
        for i, (label, intentados, completados, perdidos, pct) in enumerate(zonas_vals):
            pct_total = intentados / total_intentados * 100
            size_px = 55 + int((intentados / max_int) * 85)
            bubble_sizes.append(size_px)
            bubble_texts.append(str(intentados))
            # % total arriba del campo
            fig_cp.add_annotation(x=zone_cx[i], y=pitch_h+5,
                text=f"<b>{pct_total:.0f}%</b>", showarrow=False,
                font=dict(color=AZUL_CELESTE, size=12, family="Inter"),
                xref="x", yref="y")

        fig_cp.add_trace(go.Scatter(
            x=zone_cx, y=[pitch_h/2]*3,
            mode="markers+text",
            marker=dict(size=bubble_sizes, sizemode="diameter",
                        color="rgba(15,25,50,0.88)",
                        line=dict(color=AZUL_CELESTE, width=2.5)),
            text=bubble_texts,
            textposition="middle center",
            textfont=dict(color=BLANCO, size=20, family="Inter", weight=700),
            showlegend=False, hoverinfo="skip",
        ), row=1, col=1)

        # Flecha dirección ataque
        fig_cp.add_annotation(x=10, y=6, ax=2, ay=6,
            xref="x", yref="y", axref="x", ayref="y",
            text="", showarrow=True, arrowhead=2, arrowsize=1.2,
            arrowwidth=2, arrowcolor="rgba(255,255,255,0.4)")

        # ── Row 2: barras apiladas ──
        completados_vals = [v[2] for v in zonas_vals]
        perdidos_vals    = [v[3] for v in zonas_vals]
        pct_vals         = [v[4] for v in zonas_vals]
        max_bar = max(c+p for c, p in zip(completados_vals, perdidos_vals)) or 1

        fig_cp.add_trace(go.Bar(
            name="Pases completados", x=zone_cx, y=completados_vals,
            marker=dict(color=AZUL_CELESTE, line=dict(color=AZUL_OSCURO, width=1)),
            text=completados_vals, textposition="inside",
            textfont=dict(color=BLANCO, size=13, family="Inter"),
            insidetextanchor="middle", width=22,
        ), row=2, col=1)
        fig_cp.add_trace(go.Bar(
            name="Pases perdidos", x=zone_cx, y=perdidos_vals,
            marker=dict(color=GRIS_MEDIO, opacity=0.55,
                        line=dict(color=AZUL_OSCURO, width=1)),
            text=perdidos_vals, textposition="inside",
            textfont=dict(color=AZUL_OSCURO, size=12, family="Inter"),
            insidetextanchor="middle", width=22,
        ), row=2, col=1)

        # % bajo las barras y etiquetas DEF/MED/ATQ
        for i, (label, pct) in enumerate(zip(["DEF","MED","ATQ"], pct_vals)):
            pct_color = VERDE if pct >= 80 else (DORADO if pct >= 70 else ROJO)
            fig_cp.add_annotation(x=zone_cx[i], y=-max_bar*0.12,
                text=f"<b>{pct:.0f}%</b>", showarrow=False,
                font=dict(color=pct_color, size=13, family="Inter"),
                xref="x2", yref="y2")
            fig_cp.add_annotation(x=zone_cx[i], y=-max_bar*0.28,
                text=f"<b>{label}</b>", showarrow=False,
                font=dict(color=GRIS_MEDIO, size=11, family="Inter"),
                xref="x2", yref="y2")

        # ── Layout global ──
        fig_cp.update_layout(
            paper_bgcolor=PAPER_BG,
            height=540,
            margin=dict(t=30, b=50, l=10, r=10),
            font=dict(family="Inter, Arial, sans-serif", color=BLANCO),
            barmode="stack",
            showlegend=True,
            legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=BLANCO, size=11),
                        orientation="h", y=-0.08, x=0.25),
            shapes=pitch_shapes,
            # Eje X compartido (coordenadas del campo 0-105)
            xaxis=dict(range=[-3, pitch_w+3], showgrid=False, zeroline=False,
                       showticklabels=False, fixedrange=True),
            xaxis2=dict(range=[-3, pitch_w+3], showgrid=False, zeroline=False,
                        showticklabels=False, fixedrange=True),
            # Eje Y fila 1: campo
            yaxis=dict(range=[-5, pitch_h+12], showgrid=False, zeroline=False,
                       showticklabels=False, fixedrange=True),
            # Eje Y fila 2: barras
            yaxis2=dict(range=[-max_bar*0.35, max_bar*1.15],
                        showgrid=False, zeroline=False,
                        showticklabels=False, fixedrange=True),
            plot_bgcolor=PAPER_BG,
        )
        st.plotly_chart(fig_cp, use_container_width=True,
                        config={"displayModeBar": False})

    # ── Cadenas de pase ──
    section("Cadenas de pase")

    c_total = int(temp_row.get("cadenas_total") or 0)
    c_3_5   = int(temp_row.get("cadenas_3_5")   or 0)
    c_6mas  = int(temp_row.get("cadenas_6mas")   or 0)
    c_prom  = temp_row.get("cadenas_prom") or 0
    c_max   = int(temp_row.get("cadena_max")     or 0)

    # ── Gráfico de barras cadenas ──
    fig_cad = go.Figure()

    categorias = ["3-5 pases", "6+ pases"]
    valores    = [c_3_5, c_6mas]
    colores    = [AZUL_CELESTE, DORADO]

    fig_cad.add_trace(go.Bar(
        x=categorias, y=valores,
        marker=dict(
            color=colores,
            line=dict(color=AZUL_OSCURO, width=1),
        ),
        text=valores,
        textposition="outside",
        textfont=dict(color=BLANCO, size=14, family="Inter"),
        width=0.45,
        showlegend=False,
    ))

    # Línea de promedio
    if c_prom:
        # Mapear el promedio al eje X según en qué categoría cae
        fig_cad.add_hline(
            y=c_total / 2 if c_total else 0,
            line_color="rgba(255,255,255,0.0)",  # invisible, solo para referencia
        )
        # Anotación del promedio encima del gráfico
        fig_cad.add_annotation(
            x=0.5, y=1.08, xref="paper", yref="paper",
            text=f"Promedio cadena: <b>{c_prom:.1f} pases</b>   |   Cadena más larga: <b>{c_max} pases</b>",
            showarrow=False,
            font=dict(color=GRIS_MEDIO, size=11, family="Inter"),
            align="center",
        )

    t(fig_cad, height=300, margin=dict(t=50, b=20, l=40, r=20),
      yaxis=dict(title="Nº cadenas", gridcolor="rgba(200,214,229,0.07)",
                 tickfont=dict(color=GRIS_MEDIO)))
    chart(fig_cad)

    # ── Desglose en tarjetas ──
    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    items = [
        ("Cadenas de pase totales", c_total,          AZUL_CELESTE),
        ("3 a 5 pases",             c_3_5,            AZUL_CELESTE),
        ("6 o más pases",           c_6mas,           DORADO),
        ("Promedio de cadena",       f"{c_prom:.1f}",  BLANCO),
        ("Cadena más larga",         c_max,            VERDE),
    ]

    for label, valor, color in items:
        st.markdown(f"""
        <div style="display:flex;justify-content:space-between;align-items:center;
                    padding:11px 20px;border-bottom:1px solid rgba(200,214,229,0.1);
                    background:rgba(46,75,122,0.3);border-radius:6px;margin-bottom:3px;">
            <span style="color:{GRIS_MEDIO};font-size:0.9rem;">{label}</span>
            <span style="color:{color};font-size:1.1rem;font-weight:700;">{valor}</span>
        </div>""", unsafe_allow_html=True)

    # ── Defensa & Portero ──
    section("Defensa & Portero")

    paradas   = int(temp_row.get("paradas") or 0)
    tap       = int(temp_row.get("tap")     or 0)
    ge        = int(temp_row.get("goles_enc") or 0)
    pa0       = int(temp_row.get("pa0")     or 0)

    # Tarjetas resumen defensivas
    d_cols = st.columns(4)
    for col, (lbl, val, suf, col_val) in zip(d_cols, [
        ("Paradas portero",     paradas, "",  AZUL_CELESTE),
        ("Disparos bloqueados", tap,     "",  DORADO),
        ("Goles encajados",     ge,      "",  ROJO),
        ("Portería a cero",     "✅" if pa0 else "❌", "", VERDE if pa0 else ROJO),
    ]):
        col.markdown(f"""
        <div style="background:{AZUL_MEDIO};border-radius:10px;padding:14px 16px;
                    border-top:3px solid {col_val};text-align:center;margin-bottom:8px;">
            <div style="color:{GRIS_MEDIO};font-size:0.68rem;text-transform:uppercase;
                        letter-spacing:0.08em;margin-bottom:6px;">{lbl}</div>
            <div style="color:{col_val};font-size:1.8rem;font-weight:800;">{val}</div>
        </div>""", unsafe_allow_html=True)

    # Paradas + TAP por tramo
    tramos_def = tramos.copy()
    par_vals = tramos_def["paradas"].fillna(0).tolist()
    tap_vals = tramos_def["tap"].fillna(0).tolist()
    tr_labels = tramos_def["tramo_raw"].tolist()

    if any(v > 0 for v in par_vals + tap_vals):
        fig_def = go.Figure()
        fig_def.add_trace(go.Bar(
            name="Paradas portero", x=tr_labels, y=par_vals,
            marker=dict(color=AZUL_CELESTE, line=dict(color=AZUL_OSCURO, width=1)),
            text=[str(int(v)) if v > 0 else "" for v in par_vals],
            textposition="outside", textfont=dict(color=AZUL_CELESTE, size=11),
        ))
        fig_def.add_trace(go.Bar(
            name="Disparos bloqueados (TAP)", x=tr_labels, y=tap_vals,
            marker=dict(color=DORADO, opacity=0.85, line=dict(color=AZUL_OSCURO, width=1)),
            text=[str(int(v)) if v > 0 else "" for v in tap_vals],
            textposition="outside", textfont=dict(color=DORADO, size=11),
        ))
        # Separador descanso
        if "45+" in tr_labels and "46-60" in tr_labels:
            fig_def.add_vline(
                x=tr_labels.index("45+") + 0.5,
                line_color="rgba(255,255,255,0.2)", line_width=1.5, line_dash="dash",
            )
        t(fig_def, height=280, barmode="group", bargap=0.25,
          margin=dict(t=20, b=20, l=30, r=20))
        chart(fig_def)

    # ── Jugadas ensayadas ──
    section("Jugadas ensayadas")

    corners = int(temp_row.get("jugadas_set") or 0)
    faltas  = int(temp_row.get("faltas")       or 0)
    penaltis = int(temp_row.get("pl")          or 0)
    centros_val = int(temp_row.get("centros")  or 0)

    je_cols = st.columns(4)
    for col, (lbl, val, color) in zip(je_cols, [
        ("Saques de esquina", corners,     AZUL_CELESTE),
        ("Faltas cometidas",  faltas,      ROJO),
        ("Penaltis",          penaltis,    DORADO),
        ("Centros",           centros_val, GRIS_MEDIO),
    ]):
        col.markdown(f"""
        <div style="background:{AZUL_MEDIO};border-radius:10px;padding:14px 16px;
                    border-left:4px solid {color};text-align:center;margin-bottom:8px;">
            <div style="color:{GRIS_MEDIO};font-size:0.68rem;text-transform:uppercase;
                        letter-spacing:0.08em;margin-bottom:6px;">{lbl}</div>
            <div style="color:{color};font-size:1.8rem;font-weight:800;">{val}</div>
        </div>""", unsafe_allow_html=True)

    # Faltas por tramo
    faltas_tramo = tramos["faltas"].fillna(0).tolist()
    if any(v > 0 for v in faltas_tramo):
        fig_fal = go.Figure()
        fig_fal.add_trace(go.Bar(
            name="Faltas", x=tr_labels, y=faltas_tramo,
            marker=dict(color=ROJO, opacity=0.8, line=dict(color=AZUL_OSCURO, width=1)),
            text=[str(int(v)) if v > 0 else "" for v in faltas_tramo],
            textposition="outside", textfont=dict(color=ROJO, size=11),
        ))
        fig_fal.add_hline(y=sum(faltas_tramo)/len([v for v in faltas_tramo if v > 0] or [1]),
                          line_dash="dot", line_color="rgba(255,255,255,0.3)",
                          annotation_text="Media", annotation_position="right",
                          annotation_font=dict(color=GRIS_MEDIO, size=9))
        t(fig_fal, height=240, margin=dict(t=20, b=20, l=30, r=50))
        chart(fig_fal)


# ── Página: Rivales ───────────────────────────────────────────────────────────
def page_rivales(df: pd.DataFrame):
    st.markdown(f"<h2 style='color:{BLANCO};margin-bottom:4px;'>⚔️ Rivales</h2>",
                unsafe_allow_html=True)

    temporada = df[df["nivel"] == "temporada"]
    rivales   = temporada["rival"].unique().tolist()
    rival_sel = st.selectbox("", rivales, format_func=lambda r: f"CAdF vs {r}",
                              label_visibility="collapsed")

    rival_df  = df[df["rival"] == rival_sel]
    temp_rows = rival_df[rival_df["nivel"] == "temporada"]
    tramos_df = rival_df[rival_df["nivel"] == "tramo"].copy()

    st.markdown(f"""
    <div style="background:{AZUL_MEDIO};border-radius:10px;padding:16px 20px;
                border-left:4px solid {DORADO};margin-bottom:20px;">
        <span style="color:{GRIS_MEDIO};font-size:0.72rem;text-transform:uppercase;">
            Enfrentamientos vs</span>
        <h3 style="color:{BLANCO};margin:4px 0 2px;">{rival_sel}</h3>
        <span style="color:{AZUL_CELESTE};font-size:0.85rem;">
            {len(temp_rows)} partido(s) en el histórico</span>
    </div>""", unsafe_allow_html=True)

    cols = st.columns(4)
    for col, (lbl, val, suf) in zip(cols, [
        ("Goles media",    f"{temp_rows['goles'].mean():.1f}",        ""),
        ("Posesión media", f"{temp_rows['pct_posesion'].mean():.1f}", "%"),
        ("Pases % media",  f"{temp_rows['pct_pases'].mean():.1f}",   "%"),
        ("Tiros puerta",   f"{temp_rows['tiros_puerta'].mean():.1f}", ""),
    ]):
        col.markdown(metric_card(lbl, val, suf), unsafe_allow_html=True)

    section("¿En qué momentos dominas y dónde pierdes el control?")
    tramos_mean = (
        tramos_df.groupby("tramo_raw")[["pct_posesion", "pct_pases"]]
        .mean().reindex(TRAMO_ORDER).reset_index()
    )

    fig_r = go.Figure()
    fig_r.add_hrect(y0=50, y1=100, fillcolor="rgba(106,175,230,0.05)", line_width=0)
    fig_r.add_hrect(y0=0,  y1=50,  fillcolor="rgba(232,93,117,0.05)",  line_width=0)
    fig_r.add_hline(y=50, line_dash="dot",
                    line_color="rgba(200,214,229,0.35)", line_width=1.5)

    fig_r.add_trace(go.Scatter(
        x=tramos_mean["tramo_raw"], y=tramos_mean["pct_posesion"],
        name="Posesión %", mode="lines+markers",
        line=dict(color=AZUL_CELESTE, width=3, shape="spline"),
        marker=dict(size=11, color=AZUL_CELESTE, line=dict(color=BLANCO, width=2)),
        fill="tozeroy", fillcolor="rgba(106,175,230,0.12)",
    ))
    fig_r.add_trace(go.Scatter(
        x=tramos_mean["tramo_raw"], y=tramos_mean["pct_pases"],
        name="Pases %", mode="lines+markers",
        line=dict(color=DORADO, width=2.5, dash="dash", shape="spline"),
        marker=dict(size=9, color=DORADO),
    ))

    for _, row in tramos_mean.iterrows():
        if pd.notna(row["pct_posesion"]) and row["pct_posesion"] < 45:
            fig_r.add_vrect(x0=row["tramo_raw"], x1=row["tramo_raw"],
                            fillcolor=ROJO, opacity=0.12, line_width=0)
            fig_r.add_annotation(x=row["tramo_raw"], y=22,
                                  text="⚠️", showarrow=False, font=dict(size=14))

    fig_r.update_layout(
        paper_bgcolor=AZUL_OSCURO, plot_bgcolor=PLOT_BG,
        height=400, margin=dict(t=20, b=40, l=40, r=20),
        font=dict(family="Inter, Arial, sans-serif", color=BLANCO),
        legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color=BLANCO),
                    orientation="h", y=1.08),
        yaxis=dict(range=[20, 110], ticksuffix="%",
                   gridcolor="rgba(200,214,229,0.08)",
                   tickfont=dict(color=BLANCO)),
        xaxis=dict(gridcolor="rgba(200,214,229,0.08)",
                   tickfont=dict(color=BLANCO, size=11)),
    )
    chart(fig_r)

    st.markdown(f"""
    <div style="background:{AZUL_MEDIO};border-radius:8px;padding:12px 16px;
                border-left:3px solid {ROJO};font-size:0.82rem;color:{GRIS_MEDIO};
                margin-top:12px;">
        ⚠️ Los tramos marcados en rojo son momentos donde históricamente perdéis
        el control ante <b style="color:{BLANCO}">{rival_sel}</b>.
        Útiles para preparar tácticamente el próximo encuentro.
    </div>""", unsafe_allow_html=True)


# ── GPS loader ───────────────────────────────────────────────────────────────
def load_gps() -> pd.DataFrame:
    """Carga todos los Training Report .xlsx de data/GPS/ y devuelve _synced_data."""
    import openpyxl
    gps_dir = os.path.join(os.path.dirname(__file__), "data", "GPS")
    frames = []
    if not os.path.isdir(gps_dir):
        return pd.DataFrame()
    for fname in sorted(os.listdir(gps_dir)):
        if not fname.endswith(".xlsx") or fname.startswith("~$"):
            continue
        wb = openpyxl.load_workbook(os.path.join(gps_dir, fname), data_only=True, read_only=True)
        if "_synced_data" not in wb.sheetnames:
            continue
        ws = wb["_synced_data"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        data = [dict(zip(headers, r)) for r in rows[1:]]
        frames.append(pd.DataFrame(data))
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    df = df.rename(columns={
        "Date": "fecha",
        "Player Name": "jugador",
        "Distance": "distancia",
        "GPS Load": "carga",
        "Top Speed": "vel_max",
        "Peak Accel": "accel_max",
        "Accel Zones  Count": "accel_count",
        "Speed Zones  Distance": "hsr_dist",
        "Speed Zones  Count": "hsr_count",
    })
    for col in ["distancia", "carga", "vel_max", "accel_max", "accel_count", "hsr_dist", "hsr_count"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    if "fecha" in df.columns:
        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
        df = df.dropna(subset=["fecha"])
        df["fecha_str"] = df["fecha"].dt.strftime("%d/%m/%Y")
    return df.sort_values("fecha").reset_index(drop=True)


# ── Página: Física (GPS) ──────────────────────────────────────────────────────
METRICAS_INFO = {
    "distancia": ("Distancia (km)",        " km",   "Kilómetros recorridos durante la sesión."),
    "vel_max":   ("Vel. Máxima (km/h)",    " km/h", "Velocidad punta máxima registrada."),
    "carga":     ("Carga GPS",             "",      "Índice de estrés físico total (distancia + intensidad + aceleraciones)."),
    "accel_max": ("Accel. Máxima (m/s²)",  " m/s²", "Pico de aceleración. Refleja explosividad del esfuerzo."),
    "hsr_dist":  ("HSR Distancia (km)",    " km",   "Distancia recorrida en zonas de alta velocidad."),
    "hsr_count": ("HSR Entradas",          "",      "Número de esfuerzos en alta velocidad durante la sesión."),
}
METRICA_MAP = {
    "Distancia (km)":       "distancia",
    "Carga GPS":            "carga",
    "Velocidad Máxima":     "vel_max",
    "Aceleración Máxima":   "accel_max",
    "HSR Distancia":        "hsr_dist",
    "HSR Entradas":         "hsr_count",
}

def _fmt(val, col):
    if not pd.notna(val):
        return "—"
    return f"{val:.2f}" if col in ("distancia", "accel_max") else f"{val:.1f}"

def _kpis_equipo(df_sesion):
    cols = st.columns(6)
    for col_ui, (col_key, agg) in zip(
        cols,
        [("distancia","mean"),("vel_max","max"),("carga","mean"),
         ("accel_max","mean"),("hsr_dist","mean"),("hsr_count","mean")]
    ):
        label, suffix, tooltip = METRICAS_INFO[col_key]
        val = df_sesion[col_key].agg(agg) if col_key in df_sesion.columns else None
        with col_ui:
            st.markdown(metric_card(label, _fmt(val, col_key), suffix), unsafe_allow_html=True)
            st.markdown(
                f"<div style='font-size:0.72rem;color:{GRIS_MEDIO};margin-top:4px;"
                f"line-height:1.4;padding:0 4px;'>{tooltip}</div>",
                unsafe_allow_html=True,
            )

def _evo_chart(df_evo, metrica_col, metrica_label, titulo):
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df_evo["fecha_str"], y=df_evo[metrica_col],
        name=metrica_label,
        marker=dict(color=AZUL_CELESTE, opacity=0.85, line=dict(color=DORADO, width=1)),
        text=[_fmt(v, metrica_col) for v in df_evo[metrica_col]],
        textposition="outside", textfont=dict(color=BLANCO, size=11),
    ))
    media = df_evo[metrica_col].mean()
    fig.add_hline(y=media, line_dash="dash", line_color=DORADO, line_width=2,
                  annotation_text=f"  Media: {_fmt(media, metrica_col)}",
                  annotation_font_color=DORADO, annotation_font_size=13,
                  annotation_position="top left")
    t(fig, height=360)
    fig.update_layout(title=dict(text=titulo, font=dict(color=GRIS_MEDIO, size=13)))
    chart(fig)

def page_fisica():
    st.markdown(f"""
    <h1 style="color:{BLANCO};font-size:2rem;font-weight:700;margin-bottom:2px;">
        Preparación Física
    </h1>
    <p style="color:{AZUL_CELESTE};font-size:0.9rem;margin-bottom:20px;">
        Datos GPS de entrenamiento · Titan
    </p>""", unsafe_allow_html=True)

    df = load_gps()
    if df.empty:
        st.info("No hay archivos GPS en data/GPS/. Ejecuta script.py para descargar el Training Report.")
        return

    jugadores  = sorted(df["jugador"].dropna().unique())
    fechas_str = sorted(df["fecha_str"].unique())

    # ── Filtros globales ──────────────────────────────────────────────────────
    col_f1, col_f2, col_f3 = st.columns([2, 2, 2])
    with col_f1:
        metrica_label = st.selectbox("Métrica", list(METRICA_MAP.keys()), key="gps_metrica")
        metrica_col   = METRICA_MAP[metrica_label]
    with col_f2:
        fecha_sel = st.selectbox("Sesión", ["Última"] + fechas_str, key="gps_fecha")
    with col_f3:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)

    if fecha_sel == "Última":
        df_sesion = df[df["fecha"] == df["fecha"].max()]
        fecha_label = df["fecha"].max().strftime("%d/%m/%Y")
    else:
        df_sesion   = df[df["fecha_str"] == fecha_sel]
        fecha_label = fecha_sel

    # ── Tabs ──────────────────────────────────────────────────────────────────
    tab_equipo, tab_individual, tab_glosario = st.tabs(["🏟️  Equipo", "👤  Individual", "📖  Glosario"])

    # ════════════════════════════════════════════════════════
    # TAB EQUIPO
    # ════════════════════════════════════════════════════════
    with tab_equipo:
        section(f"Resumen del equipo · {fecha_label}")
        _kpis_equipo(df_sesion)

        st.markdown("<br>", unsafe_allow_html=True)
        section(f"Evolución del equipo · {metrica_label}")
        df_evo_eq = (
            df.groupby(["fecha_str", "fecha"])[metrica_col]
            .mean().reset_index().sort_values("fecha")
        )
        _evo_chart(df_evo_eq, metrica_col, metrica_label, f"Media del equipo · {metrica_label}")

        section(f"Comparativa jugadores · {metrica_label} — {fecha_label}")
        df_comp = df_sesion.dropna(subset=[metrica_col]).sort_values(metrica_col, ascending=True)
        if df_comp.empty:
            st.info("Sin datos para esta sesión.")
        else:
            fig_comp = go.Figure(go.Bar(
                x=df_comp[metrica_col], y=df_comp["jugador"],
                orientation="h",
                marker=dict(color=AZUL_CELESTE, opacity=0.9,
                            line=dict(color="rgba(0,0,0,0.2)", width=1)),
                text=[_fmt(v, metrica_col) for v in df_comp[metrica_col]],
                textposition="outside", textfont=dict(color=BLANCO, size=11),
            ))
            t(fig_comp, height=max(300, len(df_comp) * 38))
            chart(fig_comp)

        section("Tabla de sesión")
        df_t = df_sesion[["jugador","distancia","vel_max","carga","accel_max"]].copy()
        df_t = df_t.rename(columns={
            "jugador":"Jugador","distancia":"Dist (km)",
            "vel_max":"Vel Máx","carga":"Carga GPS","accel_max":"Accel Máx"
        }).sort_values("Jugador")
        st.dataframe(df_t, use_container_width=True, hide_index=True)

    # ════════════════════════════════════════════════════════
    # TAB INDIVIDUAL
    # ════════════════════════════════════════════════════════
    with tab_individual:
        jugador_sel = st.selectbox("Jugador", jugadores, key="gps_jugador_ind")
        df_jug = df[df["jugador"] == jugador_sel].sort_values("fecha")

        if df_jug.empty:
            st.info(f"Sin datos para {jugador_sel}.")
        else:
            # KPIs: última sesión vs media temporada
            section(f"{jugador_sel} · Última sesión vs media temporada")
            ultima = df_jug[df_jug["fecha"] == df_jug["fecha"].max()].iloc[0]
            cols_ind = st.columns(6)
            for col_ui, col_key in zip(cols_ind,
                                        ["distancia","vel_max","carga","accel_max","hsr_dist","hsr_count"]):
                label, suffix, tooltip = METRICAS_INFO[col_key]
                val_ult  = ultima[col_key]  if col_key in ultima.index else None
                val_med  = df_jug[col_key].mean() if col_key in df_jug.columns else None
                delta    = (val_ult - val_med) if pd.notna(val_ult) and pd.notna(val_med) else None
                delta_color = VERDE if delta and delta >= 0 else ROJO
                delta_str   = (f"<span style='color:{delta_color};font-size:0.8rem;'>"
                               f"{'▲' if delta>=0 else '▼'} {abs(delta):.2f} vs media"
                               f"</span>") if delta is not None else ""
                with col_ui:
                    st.markdown(
                        f"<div style='background:{AZUL_MEDIO};border-radius:10px;"
                        f"padding:14px 16px;border-left:4px solid {DORADO};margin-bottom:8px;'>"
                        f"<div style='color:{GRIS_MEDIO};font-size:0.7rem;text-transform:uppercase;"
                        f"letter-spacing:0.08em;margin-bottom:4px;'>{label}</div>"
                        f"<div style='color:{BLANCO};font-size:1.8rem;font-weight:700;line-height:1;'>"
                        f"{_fmt(val_ult, col_key)}<span style='font-size:0.9rem;color:{AZUL_CELESTE}'>{suffix}</span></div>"
                        f"<div style='margin-top:6px;'>{delta_str}</div>"
                        f"<div style='font-size:0.68rem;color:{GRIS_MEDIO};margin-top:6px;line-height:1.3;'>{tooltip}</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

            # Evolución individual
            section(f"Evolución · {jugador_sel} · {metrica_label}")
            _evo_chart(df_jug, metrica_col, metrica_label, f"{jugador_sel} · {metrica_label}")

            # Tabla individual completa
            section("Historial de sesiones")
            df_hist = df_jug[["fecha_str","distancia","vel_max","carga","accel_max"]].copy()
            df_hist = df_hist.rename(columns={
                "fecha_str":"Fecha","distancia":"Dist (km)",
                "vel_max":"Vel Máx","carga":"Carga GPS","accel_max":"Accel Máx"
            })
            st.dataframe(df_hist, use_container_width=True, hide_index=True)

    # ════════════════════════════════════════════════════════
    # TAB GLOSARIO
    # ════════════════════════════════════════════════════════
    with tab_glosario:
        st.markdown(f"""
        <p style="color:{GRIS_MEDIO};font-size:0.88rem;margin-bottom:20px;">
        Métricas exportadas por Titan en el Training Report.
        Las métricas <b style="color:{BLANCO}">en negrita</b> son las que se utilizan en la App.
        </p>""", unsafe_allow_html=True)

        glosario = [
            ("Date",                    True,  "Fecha",                    "Fecha de la sesión de entrenamiento o partido."),
            ("Player Name",             True,  "Nombre del jugador",       "Nombre completo del jugador registrado en Titan."),
            ("Distance",                True,  "Distancia (km)",           "Kilómetros totales recorridos por el jugador durante la sesión."),
            ("GPS Load",                True,  "Carga GPS",                "Índice de estrés físico total calculado por Titan. Combina distancia, intensidad, aceleraciones y tiempo en alta intensidad. No tiene unidades fijas — sirve para comparar esfuerzos entre sesiones y jugadores."),
            ("Top Speed",               True,  "Velocidad Máxima (km/h)",  "Velocidad punta más alta registrada por el jugador en toda la sesión."),
            ("Speed Zones  Distance",   True,  "HSR Distancia (km)",       "Distancia recorrida a alta velocidad (≥ 21 km/h, umbral configurado en Titan). Equivalente al HSR (High Speed Running) de otros sistemas GPS. Revisar en Titan si el umbral cambia."),
            ("Speed Zones  Count",      True,  "HSR Entradas",             "Número de esfuerzos realizados a ≥ 21 km/h durante la sesión. Indica la frecuencia de acciones de alta intensidad."),
            ("Personal Bands  Distance",False, "Distancia Bandas Personales","Distancia en zonas de velocidad personalizadas por jugador según su perfil físico individual."),
            ("Personal Bands  Count",   False, "Entradas Bandas Personales","Número de entradas en las bandas de velocidad personalizadas del jugador."),
            ("Peak Accel",              True,  "Aceleración Máxima (m/s²)","Pico de aceleración más alto registrado en la sesión. Refleja la explosividad máxima del jugador."),
            ("Accel Zones  Count",      False, "Entradas Zonas Aceleración","Número total de aceleraciones que superaron los umbrales definidos en Titan."),
        ]

        # Cabecera
        st.markdown(f"""
        <div style="display:grid;grid-template-columns:1fr 1fr 2fr;gap:0;
                    background:{AZUL_CLARO};border-radius:8px 8px 0 0;padding:10px 16px;
                    font-size:0.78rem;font-weight:700;color:{BLANCO};
                    text-transform:uppercase;letter-spacing:0.06em;">
            <div>Campo en Excel</div>
            <div>Nombre en castellano</div>
            <div>Descripción</div>
        </div>""", unsafe_allow_html=True)

        for i, (campo, en_app, nombre_es, descripcion) in enumerate(glosario):
            bg = AZUL_MEDIO if i % 2 == 0 else AZUL_OSCURO
            badge = (f"<span style='background:{DORADO};color:{AZUL_OSCURO};font-size:0.65rem;"
                     f"font-weight:700;padding:2px 6px;border-radius:4px;margin-left:6px;'>"
                     f"EN APP</span>") if en_app else ""
            nombre_html = (f"<b style='color:{BLANCO}'>{nombre_es}</b>" if en_app
                           else f"<span style='color:{GRIS_MEDIO}'>{nombre_es}</span>")
            campo_html  = (f"<b style='color:{AZUL_CELESTE}'>{campo}</b>" if en_app
                           else f"<span style='color:{GRIS_MEDIO};font-size:0.85rem'>{campo}</span>")
            st.markdown(f"""
            <div style="display:grid;grid-template-columns:1fr 1fr 2fr;gap:0;
                        background:{bg};padding:12px 16px;font-size:0.85rem;
                        border-bottom:1px solid rgba(106,175,230,0.08);">
                <div>{campo_html}{badge}</div>
                <div>{nombre_html}</div>
                <div style="color:{GRIS_MEDIO};font-size:0.82rem;line-height:1.5;">{descripcion}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown(f"""
        <div style="background:{AZUL_MEDIO};border-radius:0 0 8px 8px;padding:10px 16px;
                    font-size:0.75rem;color:{GRIS_MEDIO};border-top:1px solid rgba(106,175,230,0.15);">
            Fuente: Titan by Integrated Bionics · Los datos se actualizan ejecutando <code>script.py</code>
        </div>""", unsafe_allow_html=True)


# ── CSS global ────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&display=swap');

    html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; }}
    .stApp {{ background-color: {AZUL_OSCURO}; color: {BLANCO}; }}

    /* Sidebar */
    section[data-testid="stSidebar"] {{
        background-color: {AZUL_OSCURO};
        border-right: 1px solid {AZUL_MEDIO};
    }}
    section[data-testid="stSidebar"] label {{
        color: {GRIS_MEDIO} !important;
        font-size: 0.88rem;
    }}

    /* ── Cajitas de navegación ── */
    /* Reset botón base en sidebar */
    section[data-testid="stSidebar"] .stButton button {{
        width: 100%;
        border-radius: 10px;
        padding: 12px 16px;
        font-size: 0.92rem;
        font-weight: 600;
        text-align: left;
        border: 1px solid transparent;
        transition: all 0.18s ease;
        box-shadow: none;
        letter-spacing: 0.01em;
    }}

    /* Estado INACTIVO */
    div.nav-item-inactive .stButton button {{
        background: {AZUL_MEDIO};
        color: {GRIS_MEDIO};
        border-color: rgba(106,175,230,0.1);
    }}
    div.nav-item-inactive .stButton button:hover {{
        background: {AZUL_CLARO};
        color: {BLANCO};
        border-color: rgba(106,175,230,0.3);
        transform: translateX(3px);
        box-shadow: 0 2px 8px rgba(0,0,0,0.25);
    }}

    /* Estado ACTIVO */
    div.nav-item-active .stButton button {{
        background: linear-gradient(135deg, {AZUL_CLARO} 0%, {AZUL_MEDIO} 100%);
        color: {BLANCO} !important;
        border-color: {DORADO};
        border-left: 3px solid {DORADO};
        box-shadow:
            0 2px 8px rgba(0,0,0,0.3),
            inset 0 1px 0 rgba(255,255,255,0.08);
    }}
    div.nav-item-active .stButton button:hover {{
        background: linear-gradient(135deg, {AZUL_CLARO} 0%, {AZUL_MEDIO} 100%);
        color: {BLANCO} !important;
    }}

    /* Selectbox */
    div[data-baseweb="select"] > div {{
        background-color: {AZUL_MEDIO} !important;
        border-color: {AZUL_CLARO} !important;
        color: {BLANCO} !important;
    }}

    /* ── Sombra y profundidad en gráficos Plotly ── */
    .stPlotlyChart {{
        border-radius: 14px;
        overflow: hidden;
        box-shadow:
            0 2px  4px  rgba(0, 0, 0, 0.35),
            0 6px  16px rgba(0, 0, 0, 0.30),
            0 16px 40px rgba(0, 0, 0, 0.20),
            inset 0 1px 0 rgba(255,255,255,0.04);
        border: 1px solid rgba(106,175,230,0.10);
        transition: box-shadow 0.2s ease;
    }}
    .stPlotlyChart:hover {{
        box-shadow:
            0 2px  4px  rgba(0, 0, 0, 0.40),
            0 8px  24px rgba(0, 0, 0, 0.35),
            0 20px 50px rgba(0, 0, 0, 0.25),
            inset 0 1px 0 rgba(255,255,255,0.06);
        border-color: rgba(106,175,230,0.20);
    }}

    /* ── Sombra en tarjetas HTML (metric cards, tabla resultados, etc.) ── */
    .stMarkdown div[style*="border-radius"] {{
        transition: box-shadow 0.2s ease;
    }}

    /* Radio buttons */
    .stRadio > div {{
        background: {AZUL_MEDIO};
        border-radius: 8px;
        padding: 8px 12px;
        border: 1px solid rgba(106,175,230,0.15);
    }}
    .stRadio label {{ color: {GRIS_MEDIO} !important; }}

    /* Spinner */
    .stSpinner > div {{ border-top-color: {AZUL_CELESTE} !important; }}

    /* Scrollbar */
    ::-webkit-scrollbar {{ width: 6px; height: 6px; }}
    ::-webkit-scrollbar-track {{ background: {AZUL_OSCURO}; }}
    ::-webkit-scrollbar-thumb {{
        background: {AZUL_MEDIO};
        border-radius: 3px;
    }}
    ::-webkit-scrollbar-thumb:hover {{ background: {AZUL_CLARO}; }}
    </style>
    """, unsafe_allow_html=True)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Club Argentino · Análisis",
        page_icon="🔵",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    inject_css()

    with st.spinner("Cargando datos..."):
        df = load_all()

    if df.empty:
        st.error("No se encontraron datos en /data")
        return

    page = render_sidebar()

    if   page == "Inicio":    page_inicio(df)
    elif page == "Temporada": page_temporada(df)
    elif page == "Partido":   page_partido(df)
    elif page == "Rivales":   page_rivales(df)
    elif page == "Física":    page_fisica()


if __name__ == "__main__":
    main()
